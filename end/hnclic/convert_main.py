import sys, os, zipfile,re, requests,shutil,json,glob
sys.path.append(os.path.realpath(os.curdir+"/hello_app/"))
sys.path.append(os.path.realpath(os.curdir+"/hnclic/"))
sys.path.append(os.path.realpath(os.curdir+"/data_adapter/"))
from bs4 import BeautifulSoup
import lxml
import lxml.html
import threading
import numpy as np
import pandas as pd
import pandasql,excel2img
from numpy import nan as NaN
from functools import reduce
import jinja2,datetime,chardet,time,traceback,locale,copy
from pptx import Presentation
from pptx.enum.shapes import MSO_SHAPE_TYPE#MSO_SHAPE_TYPE.EMBEDDED_OLE_OBJECT
import comtypes.client
from openpyxl.formula.translate import Translator
from openpyxl  import load_workbook
import asyncio
import aiohttp
import yaml
from handle_file import convert_file_for_xlsx,convert_file_for_txt,convert_file_for_pptx,convert_html
from utils import get_jinja2_Environment,is_number,guess_col_names,exec_template,repeat_rearrange
import data_adapter
import glb

'''
“在Windows里，time.strftime使用C运行时的多字节字符串函数strftime，
这个函数必须先根据当前locale配置来编码格式化字符串（使用PyUnicode_EncodeLocale）。”
如果不设置好locale的话，根据默认的"C" locale，底层的wcstombs函数会使用latin-1编码（单字节编码）
来编码格式化字符串，然后导致题主提供的多字节编码的字符串在编码时出错。
'''
locale.setlocale(locale.LC_CTYPE, 'chinese')
#显示所有列
pd.set_option('display.max_columns', None)

#显示所有行
pd.set_option('display.max_rows', None)
pd.set_option('display.width', None)
pd.options.display.float_format = '{:.2f}'.format


def func_time(func):
    def inner(*args,**kw):
        start_time = time.time()
        ret=func(*args,**kw)
        end_time = time.time()
        print(str(func)+'函数运行时间为：'+str(end_time-start_time)+'s')
        return ret
    return inner

def pd_read_csv(url):
    with open(url, 'rb') as f:
        data = f.read()
        f_charInfo=chardet.detect(data)                
    try:
        return pd.read_csv(url,encoding=f_charInfo['encoding'])
    except:
        return pd.read_csv(url,encoding='gb2312')
    #return pd.read_csv(url,encoding=f_charInfo['encoding'])

def load_from_file(url,d_p):
    ret=dict()
    if(os.path.isfile(url)):
        if url.find(".xlsx")>0:
            data=pd.read_excel(url)
            data[data.columns[0]]=data[data.columns[0]].astype(str)
        else:
            with open(url, 'rb') as f:
                data = f.read()
                f_charInfo=chardet.detect(data)                
            data=pd_read_csv(url)
            data[data.columns[0]]=data[data.columns[0]].astype(str)
        if (len(d_p)==0):
            d_p.append({"t": "html","start": "1","end": "10000","columns": "auto","view_columns": "","sort": "","name": "修改这里",'data_is_json':True})
        for p in d_p:
            p['data_is_json']=True
            ret[p['name']]={'data':data,'p':p,'header':data.columns.to_list(),"data_from":{'ds':d_p}}
        return ret
    raise Exception(url +'不存在')

async def load_from_url2(data_from=None,config_data=None,upload_path=None,userid=None,user_input_form_data=None):
    '''
    data_from 某网址下的取数逻辑总汇
    user_input_form_data 实际传过来的参数
    d_p_1={'t':'json','name':'b','pattern':'#reportDivaaa1thetable tr','start':'reportDivaaa1_data={"rows":','end':"/*-end-*/"}
    d_p_2={'t':'html','name':'a','pattern':'#reportDivmainthetable tr','start':4,'end':18}
    针对t:html 的end,+4 表示需要4行,不带+号的4 ，表示第4行，负数表示从后面开始倒数几行，如果为空，表示全部，最多100万行
    '''
    data_from['exec_stat']="1:开始执行"
    adapter=data_adapter.get(data_from,userid)
    await adapter.load_data_from_url(config_data.get('form_input',{}),user_input_form_data)
    #    if soup is not None and isinstance(soup,BeautifulSoup):
    #        soup.decompose()
    #        soup=None
    data_from['exec_stat']="9:成功"
    print("成功")
    ret=dict()
    for p in data_from['ds']:
        p['exec_stat']="1:开始裁剪"
        resultModel,header,data,json_props=adapter.load_data_for_p(p)
        if resultModel=="TableModel":
            data=pd.DataFrame(data,columns=header)
        elif resultModel=="JsonModel":
            data=pd.DataFrame(data)[json_props] #按json_props中指定的顺序重排
            data.columns=header#按header重设列名
            if 'id' in json_props:
                data=data.drop(['id'], axis=1)
        elif resultModel=="DataFrame":
            pass
        else:
            raise RuntimeError("适配接口只能返回TableModel或者JsonModel。请联系管理员修改程序")
        # 列名去重
        data.columns=repeat_rearrange(list(data.columns))
        header=repeat_rearrange(header)
        #删除NULL列
        if adapter.dropNaNColumn:
            data=data.replace('',NaN).dropna(axis = 1, how = "all")
            header=list(data.columns)
        #缺省列名为：s+数字
        if isinstance(p['view_columns'],str):
            data.columns=header if p['columns']=='' or p['columns'].startswith("auto") else [('s'+str(x) if is_number(x) else x) for x in range(len(data.loc[0]))]
            view_columns=str.strip(p['view_columns'])
            if view_columns!="" :
                view_set=set(view_columns.split(',') )
                data_set=set(data.columns)
                if len(view_set- data_set )>0:
                    raise RuntimeError(f"{p['name']}以下列已经被删除：{str(view_set- data_set)}。新增的列:{str(data_set - set(p['old_columns']))} " )
                data=data[[ (data.columns[int(x)] if x.isdigit() else x) for x in view_columns.split(',')]]
        if isinstance(p['view_columns'],list):
            if len(p['view_columns'])!=0 :
                data=data[p['view_columns']]

        key_column=p.get('key_column')
        if key_column is None:
            for key in data.columns :
                if str(data[key].dtype)=='object' and  len(data[key].unique())==len(data):
                    key_column=key
                    break
        if key_column is None:
            key_column=data.columns[0]
        if key_column not in data.columns:
            raise Exception(f"{p['name']}的关键字{key_column} 不在可视列表中！通常是原始报表的列名被修改了，你可以在《查看列名》的地方，将主键选为key，清空《最终数据整理》中的信息，重新执行就可以了。")
        p['key_column']=key_column
        data=data[(data[key_column]!='') & (data[key_column].isnull()==False)].reset_index(drop=True)
        data=data.replace('','None')#先删除主键为空的情况，然后其他为字符串空替换为None
        if p.get("backup",'').strip()!='':
                rptid=os.path.realpath(upload_path).split("\\")[-1]
                bak_file=os.path.realpath(os.path.join(upload_path+"../../../过往数据/", f"{rptid}_{p['name']}"))
                if os.path.exists(f"{bak_file}_上次.json"):
                    os.remove(f"{bak_file}_上次.json")
                if os.path.exists(f"{bak_file}_上次_new.json"):
                    os.rename(f"{bak_file}_上次_new.json",f"{bak_file}_上次.json")
                cur_bak_list=sorted(glob.glob(f"{bak_file}*"),key=os.path.getmtime)
                cur_bak_list=[x for x in cur_bak_list if x!=f"{bak_file}_上次_new.json"]
                if len(cur_bak_list)!=0:#只记当天的增量
                    if (datetime.datetime.now() - datetime.datetime.fromtimestamp(os.path.getmtime(cur_bak_list[-1]) )).days ==0:
                        shutil.copyfile( cur_bak_list[-1] , f"{bak_file}_上次_new.json")
                for one_file in cur_bak_list: #删除 超过2天的历史数据
                    if (datetime.datetime.now() - datetime.datetime.fromtimestamp(os.path.getmtime(one_file) )).days >2:
                        os.remove(one_file)
                #if datetime.date.today()==one['from'][2:8]:#备份当前数据
                with open(f"{bak_file}_{datetime.date.today().isoformat()}.json", 'w') as f:
                    f.write(data.to_json(orient='records',force_ascii=False))
                #data.to_csv(f"{bak_file}.csv",index=False)
        p['old_columns']=header
        p['exec_stat']="2:裁剪成功"
        ret[p['name']]={'data':data,'p':p,'header':header,'form_input':data_from['form_input'],"data_from":data_from}            
    adapter=None
    return ret

def appendData_and_execLastSql(one_ds,ret,upload_path):
    k=one_ds['name']
    v=ret[k]
    one_ds['exec_stat']="2:开始合并和执行最终sql"
    key_column=one_ds.get('key_column')
    if key_column is None:
        for key in v['data'].columns :
            if str(v['data'][key].dtype)=='object' and  len(v['data'][key].unique())==len(v['data']):
                key_column=key
                break
    if key_column is None and v['data'].empty==False:
        key_column=v['data'].columns[0]
    one_ds['key_column']=key_column
    
    #t_append=v['p'].get('append')
    #if t_append is not None  and isinstance( t_append,dict):
    #    v['p']['append']=[t_append,]
    for one in one_ds.get('append',list()):
        one_ds['exec_stat']="3:开始合并"+one['from']
        if one.get('from','')=='':
            continue
        elif one['from'].find(".xlsx")>0:
            data=pd.read_excel(os.path.join(upload_path, one['from']))
            data[data.columns[0]]=data[data.columns[0]].astype(str)
            right_key_column=data.columns[0]
        elif one['from'].find(".csv")>0:
            data=pd_read_csv(os.path.join(upload_path, one['from']))
            right_key_column=data.columns[0]
        elif one['from'][0:2] in ['上次','备份']:#备份22时05分
            other=one['from'].split(":")
            backup_name=other[1] if len(other)>1 else k
            rptid=os.path.realpath(upload_path).split("\\")[-1]
            qushu_date=datetime.date.today()+datetime.timedelta(days=-1)
            if one['from'][0:2]=='上次':
                bak_file=os.path.realpath(os.path.join(upload_path+"../../../过往数据/", f"{rptid}_{backup_name}_上次"))
            else:
                bak_file=os.path.realpath(os.path.join(upload_path+"../../../过往数据/", f"{rptid}_{backup_name}_{qushu_date.isoformat()}"))
            if os.path.exists(f"{bak_file}.json"):
                with open(f"{bak_file}.json", 'r') as f:
                    data = f.read()
                    data=pd.read_json(data)
            elif not os.path.exists(f"{bak_file}.csv"):
                data=pd.DataFrame(columns=ret[backup_name]['data'].columns)
            else:
                data=pd_read_csv(f"{bak_file}.csv")
            if data.empty:
                data=pd.DataFrame(columns=ret[backup_name]['data'].columns)
            right_key_column=ret[backup_name]['p']['key_column']
        elif ret.get(one['from']):
            data=ret[one['from']]['data']
            right_key_column=ret[one['from']]['p']['key_column']
        else:
            continue

        if right_key_column not in data.columns:
            right_key_column=data.columns[0]

        if v['data'].empty:
            v['data']=data
            key_column=right_key_column
            continue     
        data=data[(data[right_key_column]!='') & (data[right_key_column].isnull()==False)].reset_index(drop=True)
        data[right_key_column]=data[right_key_column].astype(str)
        if len(data[right_key_column].unique())!=len(data):
            raise Exception(f"数据集【{v['p']['name']}】 的合并数据集【{one['from']}】的[{right_key_column}]列数据不唯一！")

        v['data']=v['data'].merge(data,how ="left", left_on=key_column, right_on=right_key_column,suffixes=('', f"_{one['from']}")).fillna(0)
        
        
    one_ds['exec_stat']="4:合并成功，开始数据转换"
    data=v['data']
    one_ds['after_append_columns']=list(data.columns)
    if True:# v['p'].get('data_is_json',False)==False:
        start_number=False
        for x in data.columns:#尽可能的将关键字列之后的数据设置为float类型
            if x==one_ds['key_column']:
                start_number=True
                continue
            if start_number==False:
                continue
            if data[x].dtype.name=='object':
                try:
                    data[x]=data[x].astype(int)
                except:
                    try:
                        data[x]=data[x].astype(float)
                    except:
                        pass
                    pass
    one_ds['exec_stat']="5:开始执行最终sql"
    sql=one_ds.get('sql','').strip()
    if sql!="" :
        exec_sql=exec_template(None,sql,[])
        data=pandasql.sqldf(exec_sql,dict({key:value['data'] for key,value in ret.items()}))
    if(one_ds.get('vis_sql_conf') is not None and one_ds['vis_sql_conf'].get('expr','').strip()!=''):
        data=eval(k+one_ds['vis_sql_conf']['expr'],{k:data})
    v['data']=data.round(2)
    one_ds['last_columns']=data.columns.values.tolist()    
    one_ds['exec_stat']="9:完成sql执行"

@func_time
async def load_all_data(config_data,id,appendFunDict=None,upload_path=None,userid=None,user_input_form_data=None):
    print(threading.currentThread().name)
    config_data_reset_exec_stat(config_data)
    config_data['exec_stat']='1:开始载入数据'
    ret={}
    start_time = time.time()
    cur_time=time.strftime("%H时%M分")
    #取html和csv中的数据
    tasks=[]
    #loop = asyncio.get_event_loop()
    async def _inner_task(ret):
        for one in config_data['data_from']:
            if one['type']=='sql' or one['url'].startswith('结果://'):
                continue        
            elif one['type']=='file':
                filename=os.path.join(upload_path, one['url'])
                ret={**ret,**load_from_file(filename,one['ds'])}
                continue
            #elif one['type'] in ['json','html']:
            tasks.append(load_from_url2(one,config_data,upload_path,userid,user_input_form_data=user_input_form_data))
        return await asyncio.gather(*tasks,return_exceptions=True),ret
        #status_list = loop.run_until_complete(asyncio.gather(*tasks))
    #https://yanbin.blog/how-flask-work-with-asyncio/#more-10368 关于flask中的异步，这里讲的比较详细
    #status_list,ret=asyncio.run(_inner_task(ret)) 
    status_list,ret=await _inner_task(ret)
    
    for t in status_list:
        if isinstance(t,dict):
            ret={**ret,**t}  
        if isinstance(t,Exception):
            raise t
    print( f"全部取数结束，用时： {time.time()-start_time}")
    config_data['exec_stat']='1:开始计算合并数据和执行最终sql'
    #用已定义的全局参数，覆盖所有子取数的参数合集
    form_input={}
    for one in config_data['data_from']:
        form_input={**form_input,** {x['name']:x['value'] for x in one['form_input']} }
    form_input={**form_input,** {x['name']:x['value'] for x in config_data['form_input']} } 
    config_data['form_input']=[{'name':k,'value':v} for (k,v) in form_input.items()]

    config_data['exec_stat']='2:开始计算单独sql的结果集'
    #追加数据到html数据中，一般是csv或备份或其他数据集
    if config_data.get('ds_queue') is None:
        config_data['ds_queue']=[]
        for one_data_from in config_data['data_from']:
            if one_data_from['url'].startswith('结果://') or one_data_from['type']=='sql':
                continue
            for one_ds in one_data_from['ds']:
                config_data['ds_queue'].append(one_ds['name'])
        for one_data_from in config_data['data_from']:
            if one_data_from['type']=='sql':
                for one_ds in one_data_from['ds']:
                    config_data['ds_queue'].append(one_ds['name'])
        for one_data_from in config_data['data_from']:
            if one_data_from['url'].startswith('结果://'):
                for one_ds in one_data_from['ds']:
                    config_data['ds_queue'].append(one_ds['name'])

    for ds_name in config_data['ds_queue']:
        for one_data_from in config_data['data_from']:
            if one_data_from['url'].startswith('结果://'):
                continue
            for one_ds in one_data_from['ds']:
                if one_ds['name']==ds_name:
                    if one_data_from['type']=='sql':
                        ret[one_ds['name']]={'data':pd.DataFrame(),'p':one_ds}
                    appendData_and_execLastSql(one_ds,ret,upload_path)
    print(f"appendData_and_execLastSql，用时： {time.time()-start_time}")
  
    config_data['exec_stat']='4:排序'
    for k,v in ret.items():
        data=v['data']
        sort_name=str.strip(v['p'].get('sort',''))
        if sort_name!="" :
            if isinstance(sort_name,int) or is_number(str(sort_name)):
                sort_name=data.columns[int(sort_name)]
                data=data.sort_values(sort_name,ascending= False).reset_index(drop=True)
            else:
                data[[ sort_name]]=data[[sort_name]].astype(float)
                data=data.sort_values(by=sort_name,ascending= False).reset_index(drop=True)
        v['data']=data
    print(f"排序后，用时： {time.time()-start_time}")
    ds_dict={k:v['data'] for k,v in ret.items()}
    config_data['exec_stat']='5:计算变量'
    if(config_data.get("vars") is not None):#先计算所有不是依赖excel结果的变量，这样在excel中就也可以使用变量了
        for one_var in config_data["vars"]:            
            if(ds_dict.get(one_var["name"]) is not None):
                raise SyntaxError(f'变量名字<{one_var["name"]}>已被使用：')
            # 表示的就是当前变量引用的不是excel结果。
            if ds_dict.get(one_var["ds"]) is None:
                continue
            one_var['exec_stat']='1:开始计算变量'
            try:
                exec(one_var["name"]+"="+one_var["last_statement"],ds_dict)
                val= ds_dict[one_var["name"]]
                if isinstance(val , float):
                    val=round(val,2)
                    if float(val)-int(val)==0:
                        val=int(val)
                    ds_dict[one_var["name"]] =val  
            except SyntaxError as e:
                raise SyntaxError(f'变量<{one_var["name"]}>定义语法错误：'+str(e))
            except Exception as e:
                raise SyntaxError(f'变量<{one_var["name"]}>定义语法错误：'+str(e))
            one_var['exec_stat']='9:计算变量成功'
            
    print(f"变量后，用时： {time.time()-start_time}")
    ret_files=[]
    result=''
    config_data['exec_stat']='6:按模板生成结果'
    out_file=f"{upload_path}/../../tmp/{id}/"
    if os.path.exists(out_file):
        shutil.rmtree(out_file)
    os.makedirs(out_file)
    print(f"rm后，用时： {time.time()-start_time}")
    # 为了能直接引用模板结果，先按模板生成结果
    for one_part in config_data.get('template_output_act',[]):
        if one_part["canOutput"]==False or one_part["canOutput"]=="false":
            continue
        one_file=one_part["file"]
        one_part['exec_stat']='1:开始按模板生成'
        template_file=f"{upload_path}/{one_file}"            
        if not os.path.exists(template_file):
            ret_files.append({'name':one_file,'errcode':'1','message' :'无此文件，请详细检查文件名','url':''})
            continue
        out_file=f"{upload_path}/../../tmp/{id}/{one_file}"
        if(one_file[-4:]=='pptx' ):            
            loopForDS=one_part.get("loopForDS",'').strip()
            if loopForDS=='':
                convert_file_for_pptx(out_file,template_file,ds_dict)
            else:
                t_ds_dict=ds_dict.copy()
                for idx,row in ds_dict[loopForDS ].iterrows():
                    t_ds_dict["_loop_"]=row
                    t_ds_dict["_idx_"]=idx            
                    convert_file_for_pptx(out_file,template_file,t_ds_dict)
        elif(one_file[-4:]=='xlsx' ):
                convert_file_for_xlsx(out_file,template_file,ds_dict, appendFunDict=appendFunDict)
        ret_files.append({'name':one_file,'errcode':'0','message' :'成功生成','url':f'/mg/file/download_t/{id}/{one_file}'})
        one_part['exec_stat']='9:按模板生成成功'

    print(f"模板后，用时： {time.time()-start_time}")
    config_data['exec_stat']='7:从excel模板结果中取数'
    for data_from in config_data['data_from']:
        if not data_from['url'].startswith('结果://'):
            continue
        data_from['exec_stat']="1:从excel模板结果中取数开始"
        get_excel_data(data_from,id,upload_path,ds_dict,ret)
    config_data['exec_stat']='9:成功'
    return ds_dict

def get_excel_data(data_from,id,upload_path,ds_dict={},ret={}):
    ds_needInit=True if data_from['ds'] is None or len(data_from['ds'])==0 else False
    excel_result_file=f"{upload_path}/../../tmp/{id}/{data_from['url'][len('结果://'):]}"
    if os.path.exists(excel_result_file)==False:
        raise RuntimeError("【"+data_from['url'][len('结果://'):]+"】不存在，如果存在该excel模板，请先运行查看数据，然后再添加其作为数据源！")
    wb = load_workbook(excel_result_file,data_only=True)
    try:
        if len(wb.defined_names.definedName)==0:
            return
        if ds_needInit:# 没有ds定义，需要初始化ds，这里要判断是不是能不能作为ds，todo
            data_from['ds']=[{"t": "json","pattern": wb.defined_names.definedName[0].name,"end": '',
                    "start": '',"columns": "auto","view_columns": "","sort": "","name": "修改这里",
                    "old_columns":[]}]

        for one_ds in data_from['ds']:
            has_define=False
            for my_range in wb.defined_names.definedName:
                if my_range.name != one_ds['pattern']:
                    continue
                one_ds['exec_stat']='1:开始取数'
                has_define=True
                for title, coord in my_range.destinations: # returns a generator of (worksheet title, cell range) tuples
                    ws = wb[title]
                    cell_ranges=ws[coord]
                    col_nums=cell_ranges[-1][-1].column -cell_ranges[0][0].column +1 
                    row_nums=cell_ranges[-1][-1].row -cell_ranges[0][0].row +1
                    excel_results= [[None] * col_nums for i in range(row_nums)]
                    for row in cell_ranges:
                        for cell in row:
                            if cell.value is None:
                                continue
                            for one_merged_cell in ws.merged_cells:
                                if cell.row == one_merged_cell.min_row  and cell.column ==one_merged_cell.min_col :
                                    for i_row in range(one_merged_cell.max_row-one_merged_cell.min_row +1):
                                        for i_col in range(one_merged_cell.max_col-one_merged_cell.min_col+1 ):
                                            excel_results[i_row + cell.row - cell_ranges[0][0].row][i_col+cell.column -cell_ranges[0][0].column]=cell.value
                                    continue
                            excel_results[cell.row - cell_ranges[0][0].row][cell.column -cell_ranges[0][0].column]=cell.value
                    header,end_line=guess_col_names(excel_results,"auto")
                    data=pd.DataFrame(excel_results[end_line:],columns=header)
                    ret[one_ds['name']]={'data':data,'header':header,'p':one_ds}
                    
                    one_ds['last_columns']=header
                    one_ds['old_columns']=header
                    appendData_and_execLastSql(one_ds,ret,upload_path)
                    ds_dict[one_ds['name']]=ret[one_ds['name']]['data'] # 添加到变量表中
                    break
                one_ds['exec_stat']='9:成功'
            if has_define==False:
                raise Exception(data_from['url'] +'，不存在名称：'+one_ds['pattern'])
        data_from['exec_stat']="9:从excel模板结果中取数成功" 
        return ret   
    finally:
        wb.close()
        wb=None

async def files_template_exec(id,config_data,userid,app_save_path,appendFunDict=None,wx_queue=None):
    '''
    生成模板文件
    '''
    if wx_queue is None:
        wx_queue=glb.msg_queue
    upload_path=f"{app_save_path}\\{userid}\\{id}"
    config_data_reset_exec_stat(config_data)
    ds_dict=await load_all_data(config_data,id,appendFunDict,upload_path=upload_path,userid=userid)
    #ds_dict={**{k:v['data'] for k,v in ret_dataset.items()},**ds_dict}
    config_data['exec_stat']='7:计算excel模板结果中的变量'
    if(config_data.get("vars") is not None):
        for one_var in config_data["vars"]:
            # 只计算没有计算过的变量
            if(ds_dict.get(one_var["name"]) is not None):
                continue
            one_var['exec_stat']='1:开始计算变量'
            try:
                exec(one_var["name"] +"="+one_var["last_statement"],ds_dict)
                val= ds_dict[one_var["name"]]
                if isinstance(val , float):
                    val=round(val,2)
                    if float(val)-int(val)==0:
                        val=int(val)
                    ds_dict[one_var["name"]]=val
                
            except SyntaxError as e:
                raise SyntaxError(f'变量<{one_var["name"]}>定义语法错误：'+e.text)
            except Exception as e:
                raise SyntaxError(f'变量<{one_var["name"]}>定义语法错误：'+str(e))
            one_var['exec_stat']='9:计算变量成功'
    config_data['exec_stat']='8:发送结果'
    ret_files=[]
    result=''
    for one_part in config_data.get('template_output_act',[]):
        if one_part["canOutput"]==False or one_part["canOutput"]=="false":
            continue
        one_file=one_part["file"]
        template_file=f"{upload_path}/{one_file}"            
        if not os.path.exists(template_file):
            ret_files.append({'name':one_file,'errcode':'1','message' :'无此文件，请详细检查文件名','url':''})
            continue
        out_file=f"{upload_path}/../../tmp/{id}/{one_file}"
        if(one_file[-4:]=='pptx' ):
            out_file=out_file.replace("\\","/")                  
            for wx_user in one_part["wx_msg"].strip().split(",")  :
                for one_img in glob.glob(out_file[:-5]+"_*/*.JPG"):
                    one_img=one_img.replace("\\","/")
                    wx_queue.put({'type':'sendImage',"wxid":wx_user,"content":one_img})
            for wx_user in one_part["wx_file"].strip().split(","):
                wx_queue.put({'type':'sendFile',"wxid":wx_user,"content":out_file})        
        elif(one_file[-4:]=='xlsx' ):
                out_file=out_file.replace("\\","/")
                for wx_user in one_part["wx_msg"].strip().split(","):
                    for one_img in glob.glob(out_file+"*.png"):
                        one_img=one_img.replace("\\","/")
                        wx_queue.put({'type':'sendImage',"wxid":wx_user,"content":one_img})
                for wx_user in one_part["wx_file"].strip().split(","):
                    wx_queue.put({'type':'sendFile',"wxid":wx_user,"content":out_file})
        elif(one_file[-3:]=='txt' ):
                message=convert_file_for_txt(out_file,template_file,ds_dict)
                out_file=out_file.replace("\\","/")
                for wx_user in one_part["wx_file"].strip().split(","):
                    wx_queue.put({'type':'sendFile',"wxid":wx_user,"content":out_file})
                #神奇的作用，emoji可以发送到微信中了
                #message=json.dumps(message)[1:-1].encode().decode('unicode_escape') 
                for wx_user in one_part["wx_msg"].strip().split(","):
                    wx_queue.put({'type':'sendMessage',"wxid":wx_user,"content":message})
        ret_files.append({'name':one_file,'errcode':'0','message' :'成功生成','url':f'/mg/file/download_t/{id}/{one_file}'})

    tpl_results=[]

    def loop_one_txt(one_part,t_ds_dict,idx=0):
        expr_html=lxml.html.fromstring(one_part['txt']).text_content()
        if expr_html.startswith("http"):
            last_append=datetime.datetime.now().strftime("%#d%#H%#M%S")
            txt_tpl=f"http://hnapp.e-chinalife.com/weixin2/RedirctHandler2.aspx/637A7394-C8FE-4A8B-9D3A-7E7ADA492CE4/a{id}_{last_append}_{idx}.html"
            convert_html(f"{upload_path}/../../tmp/html/a{id}_{last_append}_{idx}.html",expr_html.getText(),t_ds_dict)
        else:
            txt_tpl=exec_template(None,expr_html,t_ds_dict) 
        tpl_results.append({'name': one_part['name'],"result":txt_tpl.replace('\n','\n<br>'),
            "img": 'https://gw.alipayobjects.com/zos/rmsportal/WdGqmHpayyMjiEhcKoVE.png'}
        )
        message=txt_tpl#json.dumps(txt_tpl)[1:-1].encode().decode('unicode_escape')
        for wx_user in one_part.get("wx_msg",'').strip().split(","):
            if wx_user.strip()!='':
                wx_queue.put({'type':'sendMessage',"wxid":wx_user,"content":message})

    
    for one_part in config_data.get('text_tpls',[]):
        t_ds_dict=ds_dict.copy()
        one_part['exec_stat']='1:开始按文本模板生成'
        loopForDS=one_part.get("loopForDS",'').strip()
        if loopForDS=='':
            loop_one_txt(one_part,t_ds_dict)
        else:
            for idx,row in ds_dict[loopForDS ].iterrows():
                t_ds_dict["_loop_"]=row
                t_ds_dict["_idx_"]=idx+1
                loop_one_txt(one_part,t_ds_dict,idx)
        one_part['exec_stat']='9:按文本模板生成成功'

    out_files=f"{upload_path}/../../tmp/{id}/"
    all_files=[]
    if os.path.exists(out_files):
        #out_files=os.listdir(out_files)
        for maindir, subdir, file_name_list in os.walk(out_files):
            for filename in file_name_list:
                apath = os.path.join(maindir, filename)[len(out_files):]#合并成一个完整路径
                all_files.append(apath)
           
        
    config_data['exec_stat']='9:成功'
    return ret_files,tpl_results,all_files,config_data # ,ds_dict

def config_data_reset_exec_stat(config_data):
    for data_from in config_data.get('data_from',[]):
        data_from['exec_stat']='0:未开始'
        for ds in data_from.get('ds',[]):
            ds['exec_stat']='0:未开始'
    for one in config_data.get('vars',[]):
        one['exec_stat']='0:未开始'
    for one in config_data.get('template_output_act',[]):
        one['exec_stat']='0:未开始'
    for one in config_data.get('text_tpls',[]):
        one['exec_stat']='0:未开始'


if __name__ == '__main__':
    import glb
    import objgraph  
    import gc,tracemalloc
    tracemalloc.start()
    b_snapshot = tracemalloc.take_snapshot()
    for i in range(10):
        with glb.db_connect() as conn:
            with conn.cursor(as_dict=True) as cursor:
                cursor.execute("SELECT * FROM zhanbao_tbl WHERE  id=4274 order by id asc")
                row = cursor.fetchone()
                while row:
                    b1_snapshot = tracemalloc.take_snapshot()
                    try:
                        print('worker_no:'+ row['worker_no']+"\t"+ str(row['id']) +"     "+ str(tracemalloc.get_traced_memory()))
                        files_template_exec(row['id'],json.loads(row['config_txt']),row['worker_no'],glb.config['UPLOAD_FOLDER'] ,wx_queue=glb.msg_queue)  
                        print("====================================")
                        print('worker_no:'+ row['worker_no']+"\t"+ str(row['id']) +"     "+ str(tracemalloc.get_traced_memory()))
                        print("====================================")
                        snapshot2 = tracemalloc.take_snapshot()
                        top_stats = snapshot2.compare_to(b1_snapshot, 'lineno')
                        for stat in top_stats[:10]:
                            print(stat)
                        print("====================================")
                    except Exception as e:
                        print(e)
                    row = cursor.fetchone()
        gc.collect() 
        objgraph.show_most_common_types(limit=5)   
    ### 打印出对象数目最多的 50 个类型信息  
    gc.collect()    
    #_main()
